from openpyxl import load_workbook

choice = int(input("1. Admin\n2. Passenger\nPlease choose: "))

if choice==1:    
    username = input("Enter username: ")
    password = input("Enter password: ")

    if username=="Admin" and password=="Admin":
        print("Welcome Admin!")

        choice1=1
        
        while choice1==1 or choice1==2 or choice1==3 or choice1==4 or choice1==5 or choice1==6:
            choice1 = int(input("1. Add Station\n2. Add Passenger\n3. Delete Station\n4. Delete Passenger\n5. Update Path Fare\n6. Update Passenger Balance\n7. Exit\nPlease choose: "))          
            
            path = "Data.xlsx"

            wb = load_workbook(path) 

            sheet1 = wb["Users"]
            sheet2 = wb["Stations"]

            row1 = sheet1.max_row
            row2 = sheet2.max_row
            col1 = sheet1.max_column
            col2 = sheet2.max_column 
            
            if (choice1<1) or (choice1>7):
                print("Invalid input!")
        
            if choice1==7:
                print("Bye!")
            
            if choice1==1:
                x=input("Enter station: ")
                y=0
                sheet2.cell(row=row2+1,column=1).value = x
                sheet2.cell(row=1, column=col2+1).value = x
                sheet2.cell(row=row2+1, column=col2+1).value = y
                
                for i in range(2, row2 + 1):
                    y=int(input("Enter fare till "+sheet2.cell(row=i,column=1).value+": "))
                    sheet2.cell(row=i,column=col2+1).value = y
                    sheet2.cell(row=row2+1,column=i).value = y
        
                wb.save('Data.xlsx')    
                print("Station added!")
            
            if choice1==2:
                x=input("Enter username: ")
                y=input("Enter password: ")
                z=int(input("Enter balance: "))
                
                sheet1.cell(row=row1+1,column=1).value = x
                sheet1.cell(row=row1+1, column=2).value = y
                sheet1.cell(row=row1+1, column=3).value = z
                
                wb.save('Data.xlsx')
                print("Passenger added!")
        
            if choice1==3:
                for i in range(2, row2 + 1):
                    print(i-1,". "+sheet2.cell(row=i,column=1).value)
                    
                x=int(input("Enter station to be deleted: "))

                sheet2.delete_rows(x+1, 1)
                sheet2.delete_cols(x+1, 1)
                                                
                wb.save('Data.xlsx')
                print("Station deleted!")
            
            if choice1==4:
                for i in range(2, row1 + 1):
                    print(i-1,". "+sheet1.cell(row=i,column=1).value)
                    
                x=int(input("Enter user to be deleted: "))

                sheet1.delete_rows(x+1, 1)
                                                
                wb.save('Data.xlsx')
                print("User deleted!")    
    
            if choice1==5:
                for i in range(2, row2 + 1):
                    print(i-1,". "+sheet2.cell(row=i,column=1).value)
                
                x=int(input("Enter first station of the path for which fare is to be updated: "))
                y=int(input("Enter second station of the path for which fare is to be updated: "))           
            
                z=int(input("Enter the new fare: "))
                
                sheet2.cell(x+1,y+1).value=z
                sheet2.cell(y+1,x+1).value=z
                
                wb.save('Data.xlsx')
                print("Fare updated!")
                
            if choice1==6:
                for i in range(2, row1 + 1):
                    print(i-1,". "+sheet1.cell(row=i,column=1).value)
                
                x=int(input("Enter user whose balance is to be updated: "))          
                y=int(input("Enter user amount to be added: "))
                
                sheet1.cell(x+1,3).value=sheet1.cell(x+1,3).value+y
                
                wb.save('Data.xlsx')
                print("Balance added!") 
                
    else:
        print("Invalid credentials!")
        
if choice==2:
    username = input("Enter username: ")
    password = input("Enter password: ")    
    
    path = "Data.xlsx"

    wb = load_workbook(path) 

    sheet1 = wb["Users"]
    sheet2 = wb["Stations"]

    row1 = sheet1.max_row
    row2 = sheet2.max_row
    col1 = sheet1.max_column
    col2 = sheet2.max_column
    
    for i in range(2, row1 + 1): 
        cell_uname = sheet1.cell(row = i, column = 1) 
        cell_pass = sheet1.cell(row = i, column = 2)
        cell_bal = sheet1.cell(row = i, column = 3)
        x = cell_uname.value
        y = cell_pass.value
        z = int(cell_bal.value)
    
        if x==username and y==password:            
            print("Welcome",x,"!")
            print("Your balance is Rs.",z,"/-")
            
            choice1=int(input("1. Book Ticket\n2. Exit\nPlease choose: "))
            
            if choice1==1:
                for j in range(2, row2 + 1):
                    print(j-1,". "+sheet2.cell(row=j,column=1).value)
                
                x=int(input("Enter from station: "))
                y=int(input("Enter to station: "))           
            
                z=z-sheet2.cell(x+1,y+1).value
            
                if z<0:
                    print("Low balance!")
                
                else:
                    sheet1.cell(row=i, column=3).value = z
                    wb.save('Data.xlsx')
                    print("Ticket booked successfully!")     
                    
            if choice1==2:
                print("Bye!")
            
            if (choice1<1) or (choice1>2):
                print("Invalid input!")
            
            break
            
    else:
        print("Invalid credentials!")

x=input("Press any key to close!") 